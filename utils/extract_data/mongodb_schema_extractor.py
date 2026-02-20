#!/usr/bin/env python3
"""
Extrai estrutura MongoDB: nomes de coleções e campos com tipos
IDs mostram a coleção de referência
"""

import sys
from pymongo import MongoClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from bson import ObjectId
from bson.dbref import DBRef
import re

# Configuração
HOST = sys.argv[1] if len(sys.argv) > 1 else "localhost"
PORT = int(sys.argv[2]) if len(sys.argv) > 2 else 27017
DB_NAME = sys.argv[3] if len(sys.argv) > 3 else "udata"

print(f"Conectando a {HOST}:{PORT}/{DB_NAME}...")

try:
    client = MongoClient(f"mongodb://{HOST}:{PORT}/", serverSelectionTimeoutMS=5000)
    db = client[DB_NAME]
    collections = db.list_collection_names()
    print(f"Conectado! Encontradas {len(collections)} coleções.\n")
except Exception as e:
    print(f"Erro: {e}")
    sys.exit(1)


def get_type_and_ref(key, value, all_collections):
    """Retorna o tipo e possível referência de um campo."""

    # ObjectId
    if isinstance(value, ObjectId):
        # Tentar inferir coleção pela chave
        possible_ref = infer_collection(key, all_collections)
        if possible_ref:
            return f"ObjectId (ref: {possible_ref})"
        return "ObjectId"

    # DBRef
    elif isinstance(value, DBRef):
        return f"DBRef (ref: {value.collection})"

    # String que pode ser ID
    elif isinstance(value, str):
        # Se termina em _id ou Id, pode ser referência
        if re.search(r"(_id|Id)$", key):
            possible_ref = infer_collection(key, all_collections)
            if possible_ref:
                return f"string (ref: {possible_ref})"
        return "string"

    # Integer que pode ser ID
    elif isinstance(value, int):
        if re.search(r"(_id|Id)$", key):
            possible_ref = infer_collection(key, all_collections)
            if possible_ref:
                return f"integer (ref: {possible_ref})"
        return "integer"

    # Outros tipos
    elif isinstance(value, bool):
        return "boolean"
    elif isinstance(value, float):
        return "float"
    elif isinstance(value, list):
        return "array"
    elif isinstance(value, dict):
        return "object"
    elif value is None:
        return "null"
    else:
        return type(value).__name__


def infer_collection(field_name, all_collections):
    """Tenta inferir o nome da coleção baseado no nome do campo."""
    # Remove sufixos _id, Id, _ids
    clean_name = re.sub(r"(_id|Id|_ids|Ids)$", "", field_name)

    # Tentar encontrar coleção correspondente
    for coll in all_collections:
        coll_singular = coll.rstrip("s")
        if (
            clean_name.lower() == coll.lower()
            or clean_name.lower() == coll_singular.lower()
        ):
            return coll

    return None


def extract_fields(obj, prefix="", all_collections=None):
    """Extrai campos recursivamente."""
    fields = []

    for key, value in obj.items():
        field_path = f"{prefix}.{key}" if prefix else key
        field_type = get_type_and_ref(key, value, all_collections)

        fields.append((field_path, field_type))

        # Se for objeto, processar recursivamente
        if isinstance(value, dict) and not isinstance(value, DBRef):
            fields.extend(extract_fields(value, field_path, all_collections))

        # Se for array com objetos, processar primeiro elemento
        elif isinstance(value, list) and value:
            if isinstance(value[0], dict):
                # Usar notação de array
                fields.extend(
                    extract_fields(value[0], f"{field_path}[0]", all_collections)
                )

    return fields


# Coletar estrutura de todas as coleções
all_data = {}

for coll_name in sorted(collections):
    print(f"Processando: {coll_name}")
    coll = db[coll_name]

    # Pegar primeiro documento
    doc = coll.find_one()

    if not doc:
        all_data[coll_name] = []
        continue

    # Extrair campos
    fields = extract_fields(doc, all_collections=collections)
    all_data[coll_name] = fields

# Criar Excel
print("\nCriando ficheiro Excel...")
wb = Workbook()
ws = wb.active
ws.title = "Estrutura MongoDB"

# Cabeçalhos
ws["A1"] = "Coleção"
ws["B1"] = "Campo"
ws["C1"] = "Tipo"

for cell in ["A1", "B1", "C1"]:
    ws[cell].font = Font(bold=True, color="FFFFFF")
    ws[cell].fill = PatternFill("solid", start_color="4472C4")
    ws[cell].alignment = Alignment(horizontal="center")

# Preencher dados
row = 2
for coll_name in sorted(all_data.keys()):
    fields = all_data[coll_name]

    if not fields:
        ws[f"A{row}"] = coll_name
        ws[f"B{row}"] = "(vazia)"
        ws[f"C{row}"] = "-"
        row += 1
        continue

    for field_path, field_type in fields:
        ws[f"A{row}"] = coll_name
        ws[f"B{row}"] = field_path
        ws[f"C{row}"] = field_type

        # Destacar referências
        if "ref:" in field_type:
            ws[f"C{row}"].font = Font(color="0000FF", bold=True)

        row += 1

# Ajustar larguras
ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 40
ws.column_dimensions["C"].width = 35

# Salvar
output_file = "mongodb_estrutura.xlsx"
wb.save(output_file)

print(f"\nFicheiro criado: {output_file}")
print(f"Total de coleções: {len(collections)}")
print(f"Total de campos: {row - 2}")
